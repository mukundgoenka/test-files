"""
Document Processing Service
===========================

Docling-native chunking via HybridChunker:
  - Docling DocumentConverter parses layout (headings, tables, images, lists)
  - HybridChunker respects document structure + enforces token budget
  - No LangChain dependency for chunking

Page numbers extracted from Docling provenance (PDF pages, PPTX slides, XLSX sheets).
"""
from openpyxl import load_workbook
import pandas as pd
import hashlib
import logging
import os
import re
import shutil
import tempfile
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, AsyncGenerator, Dict, List, Optional

from fastapi import UploadFile

from app.config import Settings, get_settings

logger = logging.getLogger(__name__)


# ============================================================================
# Data Classes
# ============================================================================

@dataclass
class ValidationResult:
    """Result of file validation."""
    valid: bool
    message: str
    file_name: str = ""
    file_size_mb: float = 0.0
    file_type: str = ""


@dataclass
class ChunkData:
    """Represents a document chunk with structure-aware metadata."""
    content: str
    metadata: Dict[str, Any]
    chunk_index: int
    file_name: str
    file_type: str
    content_type: str = "text"
    page_numbers: List[int] = field(default_factory=list)
    headings: List[str] = field(default_factory=list)
    heading_path: str = ""
    doc_title: str = ""
    chunk_hash: str = ""
    embed_text: str = ""
    section_name: str = ""
    caption: str = ""
    file_hash: str = ""
    is_active: bool = True
    version: int = 1


@dataclass
class ProcessedDocument:
    """Result of document processing."""
    content: str
    metadata: Dict[str, Any]
    chunks: List[ChunkData]
    file_name: str
    file_hash: str
    embedded_documents: List["ProcessedDocument"] = field(default_factory=list)
    success: bool = True
    error: Optional[str] = None


@dataclass
class ProcessingStatus:
    """Status update during processing."""
    file_name: str
    status: str  # "starting", "processing", "chunking", "complete", "error"
    progress: float  # 0.0 to 1.0
    message: str
    chunks_created: int = 0
    error: Optional[str] = None


# ============================================================================
# Document Service
# ============================================================================

_CONTENT_TYPE_MAP = {
    "text": "text",
    "table": "table",
    "picture": "image",
    "figure": "image",
    "list_item": "list",
    "caption": "text",
    "formula": "text",
    "code": "code",
    "section_header": "text",
    "page_header": "text",
    "page_footer": "text",
}


class DocumentService:
    """
    Service for document processing operations.

    Handles:
    - File validation
    - Document conversion (PDF, DOCX, PPTX, XLSX, images)
    - OCR for images and scanned PDFs
    - Structure-aware chunking via Docling HybridChunker
    - Embedded document extraction
    """

    def __init__(self, config: Optional[Settings] = None):
        self.config = config or get_settings()
        self._converter = None          # Default converter (follows global config)
        self._converter_cache = {}       # On-demand converters keyed by (ocr, table) tuple
        self._chunker = None
        self._token_encoder = None
        self._initialized = False

    def _ensure_initialized(self) -> None:
        """Lazy initialization of heavy dependencies."""
        if self._initialized:
            return

        try:
            if self.config.hf_offline_mode:
                os.environ["HF_HUB_OFFLINE"] = "1"
                os.environ["TRANSFORMERS_OFFLINE"] = "1"
                os.environ["HF_DATASETS_OFFLINE"] = "1"
                logger.info("HuggingFace offline mode enabled")

            if self.config.hf_cache_dir:
                cache_dir = self.config.hf_cache_dir
                os.environ["HF_HOME"] = cache_dir
                os.environ["HF_HUB_CACHE"] = os.path.join(cache_dir, "hub")
                os.environ["HUGGINGFACE_HUB_CACHE"] = os.path.join(cache_dir, "hub")
                os.environ["XDG_CACHE_HOME"] = cache_dir
                logger.info(f"HuggingFace cache dir: {cache_dir}")

            from docling.document_converter import DocumentConverter, PdfFormatOption
            from docling.datamodel.base_models import InputFormat
            from docling.datamodel.pipeline_options import (
                PdfPipelineOptions,
                TesseractCliOcrOptions,
            )
            from docling.backend.pypdfium2_backend import PyPdfiumDocumentBackend
            from docling.pipeline.standard_pdf_pipeline import StandardPdfPipeline
            from docling_core.transforms.chunker import HybridChunker
            from docling_core.transforms.chunker.tokenizer.openai import OpenAITokenizer
            import tiktoken

            tesseract_path = os.path.expandvars(self.config.tesseract_path)
            logger.info(f"Configuring Tesseract OCR with path: {tesseract_path}")
            ocr_options = TesseractCliOcrOptions(
                lang=["eng"],
                force_full_page_ocr=False,
                tesseract_cmd=tesseract_path,
            )

            logger.info(f"PDF OCR enabled: {self.config.ocr_enabled}, Table extraction: {self.config.table_extraction_enabled}")
            pipeline_options = PdfPipelineOptions(
                do_ocr=self.config.ocr_enabled,
                do_table_structure=self.config.table_extraction_enabled,
                ocr_options=ocr_options,
                generate_page_images=False,
                generate_picture_images=False,
                images_scale=1.0,
                force_backend_text=True,
            )

            pdf_format_option = PdfFormatOption(
                pipeline_cls=StandardPdfPipeline,
                backend=PyPdfiumDocumentBackend,
                pipeline_options=pipeline_options,
            )


            self._converter = DocumentConverter(
                allowed_formats=[
                    InputFormat.PDF,
                    InputFormat.DOCX,
                    InputFormat.PPTX,
                    InputFormat.IMAGE,
                    InputFormat.XLSX,
                    InputFormat.CSV,
                    InputFormat.MD,
                ],
                format_options={
                    InputFormat.PDF: pdf_format_option,
                }
            )

            # Store references for building per-request converters
            self._ocr_options = ocr_options
            self._InputFormat = InputFormat
            self._PdfFormatOption = PdfFormatOption
            self._PdfPipelineOptions = PdfPipelineOptions
            self._StandardPdfPipeline = StandardPdfPipeline
            self._PyPdfiumDocumentBackend = PyPdfiumDocumentBackend
            self._DocumentConverter = DocumentConverter

            enc = tiktoken.get_encoding(self.config.chunk_tokenizer)
            self._token_encoder = enc
            tokenizer = OpenAITokenizer(tokenizer=enc, max_tokens=self.config.chunk_max_tokens)
            self._chunker = HybridChunker(tokenizer=tokenizer, merge_peers=True)

            self._initialized = True
            logger.info(
                f"DocumentService initialized (Docling HybridChunker: "
                f"tokenizer={self.config.chunk_tokenizer}, "
                f"max_tokens={self.config.chunk_max_tokens})"
            )

        except ImportError as e:
            logger.error(f"Failed to import document processing dependencies: {e}")
            raise

    @property
    def is_initialized(self) -> bool:
        return self._initialized

    def validate_file(self, file: UploadFile) -> ValidationResult:
        """Validate an uploaded file (extension + size)."""
        file_name = file.filename or "unknown"
        file_ext = Path(file_name).suffix.lower()

        if file_ext not in self.config.supported_extensions:
            return ValidationResult(
                valid=False,
                message=f"Unsupported file type: {file_ext}. Supported: {', '.join(self.config.supported_extensions)}",
                file_name=file_name,
                file_type=file_ext,
            )

        if hasattr(file, 'size') and file.size:
            file_size_mb = file.size / (1024 * 1024)
            if file_size_mb > self.config.max_file_size_mb:
                return ValidationResult(
                    valid=False,
                    message=f"File too large: {file_size_mb:.2f} MB. Maximum: {self.config.max_file_size_mb} MB",
                    file_name=file_name,
                    file_size_mb=file_size_mb,
                    file_type=file_ext,
                )

        return ValidationResult(valid=True, message="File is valid", file_name=file_name, file_type=file_ext)

    def get_file_hash(self, content: bytes) -> str:
        """Generate MD5 hash for file content."""
        hasher = hashlib.md5(usedforsecurity=False)
        hasher.update(content)
        return hasher.hexdigest()

    def _get_file_hash_from_path(self, file_path: str) -> str:
        """Generate MD5 hash for file at path."""
        hasher = hashlib.md5(usedforsecurity=False)
        with open(file_path, 'rb') as f:
            buf = f.read(65536)
            while len(buf) > 0:
                hasher.update(buf)
                buf = f.read(65536)
        return hasher.hexdigest()

    async def save_upload_file(self, file: UploadFile, temp_dir: str) -> str:
        """Save uploaded file to temporary directory."""
        file_path = os.path.join(temp_dir, file.filename or "uploaded_file")
        content = await file.read()
        with open(file_path, 'wb') as f:
            f.write(content)
        await file.seek(0)
        return file_path

    def _convert_doc_to_docx(self, doc_path: str) -> Optional[str]:
        """Convert .doc file to .docx format using Microsoft Word COM automation (Windows only)."""
        try:
            import win32com.client
            import pythoncom

            # Ensure absolute path for COM
            doc_path = os.path.abspath(doc_path)
            docx_path = os.path.splitext(doc_path)[0] + '.docx'
            logger.info(f"Converting {doc_path} to {docx_path}")

            # Initialize COM for this thread
            pythoncom.CoInitialize()
            try:
                word = win32com.client.Dispatch('Word.Application')
                word.Visible = False
                word.DisplayAlerts = False

                doc = word.Documents.Open(doc_path)
                # Save as docx (FileFormat=16 is wdFormatXMLDocument/.docx)
                doc.SaveAs2(docx_path, FileFormat=16)
                doc.Close()
                word.Quit()

                logger.info(f"Successfully converted to {docx_path}")
                return docx_path
            finally:
                pythoncom.CoUninitialize()

        except ImportError:
            logger.error("pywin32 not installed. Install with: pip install pywin32")
            return None
        except Exception as e:
            logger.error(f"Failed to convert .doc to .docx: {e}")
            return None

    def extract_embedded_docx(self, docx_path: str, output_dir: str) -> List[str]:
        """Extract embedded DOCX files from a DOCX document."""
        embedded_files = []
        try:
            with zipfile.ZipFile(docx_path, 'r') as zip_file:
                for file_info in zip_file.filelist:
                    if 'word/embeddings/' in file_info.filename and file_info.filename.endswith('.docx'):
                        embedded_path = os.path.join(output_dir, os.path.basename(file_info.filename))
                        with zip_file.open(file_info) as source, open(embedded_path, 'wb') as target:
                            shutil.copyfileobj(source, target)
                        embedded_files.append(embedded_path)
                        logger.info(f"Extracted embedded DOCX: {embedded_path}")
        except Exception as e:
            logger.warning(f"Could not extract embedded files from {docx_path}: {e}")
        return embedded_files
    

     

    # ------------------------------------------------------------------
    # Chunking
    # ------------------------------------------------------------------

    def _extract_page_numbers(self, doc_items) -> List[int]:
        """
        Extract page/slide/sheet numbers from doc_item provenance.
        PDF → page_no, PPTX → slide number, XLSX → sheet index.
        """
        pages = []
        for item in doc_items:
            prov = getattr(item, "prov", None)
            if not prov:
                continue
            for p in prov:
                page_no = getattr(p, "page_no", None)
                if page_no and page_no not in pages:
                    pages.append(page_no)
        return sorted(pages)

    def _get_primary_content_type(self, doc_items) -> str:
        """Determine chunk content_type from the primary doc_item label."""
        if not doc_items:
            return "text"
        label = doc_items[0].label.value if hasattr(doc_items[0].label, 'value') else str(doc_items[0].label)
        return _CONTENT_TYPE_MAP.get(label.lower(), "text")

    def _count_tokens(self, text: str) -> int:
        """Count tokens using the configured tiktoken encoder."""
        if self._token_encoder is None:
            return len(text) // 4
        return len(self._token_encoder.encode(text))

    def _derive_table_name(self, table_item, headings: List[str], doc, fallback_index: int) -> str:
        """Derive a table name: caption -> nearest heading -> 'Table N'."""
        name = ""
        try:
            name = (table_item.caption_text(doc) or "").strip()
        except Exception:
            name = ""
        if not name and headings:
            name = headings[-1]
        if not name:
            name = f"Table {fallback_index}"
        return name

    def _format_table_chunks(self, table_item, table_name: str, doc) -> Optional[List[str]]:
        """
        Render a Docling table as header-paired row text, splitting into
        multiple row-group chunks that each fit the token budget while
        repeating the 'Table: <name>' header line.

        Returns None if the table grid is unavailable (caller falls back to
        the original markdown rendering).
        """
        try:
            df = table_item.export_to_dataframe(doc)
        except Exception:
            try:
                df = table_item.export_to_dataframe()
            except Exception:
                return None

        if df is None or df.empty:
            return None

        columns = [str(c).strip() for c in df.columns]
        header_line = f"Table: {table_name}"

        row_texts: List[str] = []
        for _, row in df.iterrows():
            cells = ["" if v is None else str(v).strip() for v in row.tolist()]
            if not any(cells):
                continue
            parts = [cells[0]] if cells else []
            for col, val in zip(columns[1:], cells[1:]):
                if val:
                    parts.append(f"{col}: {val}")
            row_texts.append("Row: " + " | ".join(parts))

        if not row_texts:
            return None

        max_tokens = self.config.chunk_max_tokens
        chunks: List[str] = []
        current: List[str] = []
        for rt in row_texts:
            trial = current + [rt]
            block = header_line + "\n" + "\n".join(trial)
            if current and self._count_tokens(block) > max_tokens:
                chunks.append(header_line + "\n" + "\n".join(current))
                current = [rt]
            else:
                current = trial
        if current:
            chunks.append(header_line + "\n" + "\n".join(current))

        return chunks or None

    def chunk_document(self, docling_doc, metadata: Dict[str, Any]) -> List[ChunkData]:
        """
        Chunk a DoclingDocument using HybridChunker.

        The chunker respects document layout (headings, tables, lists, images)
        and enforces the token budget. Post-processing builds embed_text and
        extracts page numbers from provenance.
        """
        self._ensure_initialized()
        file_name = metadata.get("file_name", "unknown")
        file_hash = metadata.get("file_hash", "")

        from docling_core.types.doc.document import TitleItem

        doc_title = ""
        try:
            for item, _level in docling_doc.iterate_items():
                if isinstance(item, TitleItem):
                    doc_title = item.text or ""
                    break
        except Exception:
            pass

        # Map self_ref -> TableItem so we can resolve table chunks back to their
        # structured grid (chunk.meta.doc_items only expose base DocItem refs).
        table_map = {
            getattr(t, "self_ref", None): t
            for t in getattr(docling_doc, "tables", []) or []
        }

        chunk_docs: List[ChunkData] = []

        for i, chunk in enumerate(self._chunker.chunk(docling_doc)):
            content = chunk.text
            if not content.strip():
                continue

            headings = chunk.meta.headings or []
            captions = chunk.meta.captions or []
            doc_items = chunk.meta.doc_items or []

            content_type = self._get_primary_content_type(doc_items)
            page_numbers = self._extract_page_numbers(doc_items)
            heading_path = " > ".join(headings) if headings else ""
            section_name = headings[-1] if headings else ""
            caption = "; ".join(captions) if captions else ""

            # Table chunks: render as header-paired rows, split big tables by
            # row-group (each chunk repeats the 'Table: <name>' header line).
            table_name = ""
            if content_type == "table":
                table_item = None
                for it in doc_items:
                    lbl = it.label.value if hasattr(it.label, "value") else str(it.label)
                    if lbl.lower() == "table":
                        table_item = table_map.get(getattr(it, "self_ref", None))
                        if table_item is not None:
                            break
                if table_item is not None:
                    table_name = self._derive_table_name(
                        table_item, headings, docling_doc, len(chunk_docs) + 1
                    )
                    formatted = self._format_table_chunks(table_item, table_name, docling_doc)
                    pieces = formatted if formatted else [content]
                else:
                    pieces = [content]
            else:
                pieces = [content]

            page_no = page_numbers[0] if page_numbers else None
            for piece in pieces:
                if not piece.strip():
                    continue

                chunk_hash = hashlib.sha256(piece.encode()).hexdigest()[:16]

                if self.config.contextualize_embeddings:
                    prefix_parts: List[str] = []
                    for part in [doc_title, *headings]:
                        if part and (not prefix_parts or prefix_parts[-1] != part):
                            prefix_parts.append(part)
                    embed_text = (" > ".join(prefix_parts) + "\n\n" + piece) if prefix_parts else piece
                else:
                    embed_text = piece

                idx = len(chunk_docs)
                meta = {
                    **metadata,
                    "content_type": content_type,
                    "page_numbers": page_numbers,
                    "page_no": page_no,
                    "headings": headings,
                    "heading_path": heading_path,
                    "section_name": section_name,
                    "doc_title": doc_title,
                    "chunk_hash": chunk_hash,
                    "chunk_size": len(piece),
                    "chunk_index": idx,
                    "total_chunks":0,
                    "is_active": True,
                    "version": 1,
                }
                if caption:
                    meta["caption"] = caption
                if table_name:
                    meta["table_name"] = table_name

                chunk_docs.append(ChunkData(
                    content=piece,
                    metadata=meta,
                    chunk_index=idx,
                    file_name=file_name,
                    file_type=metadata.get("file_type", "unknown"),
                    content_type=content_type,
                    page_numbers=page_numbers,
                    headings=headings,
                    heading_path=heading_path,
                    doc_title=doc_title,
                    chunk_hash=chunk_hash,
                    embed_text=embed_text,
                    section_name=section_name,
                    caption=caption,
                    file_hash=file_hash,
                ))

        total = len(chunk_docs)
        for idx, cd in enumerate(chunk_docs):
            cd.chunk_index = idx
            cd.metadata["chunk_index"] = idx
            cd.metadata["total_chunks"] = total

        from collections import Counter
        logger.info(
            f"Created {total} chunks for {file_name} "
            f"(title='{doc_title}', types={dict(Counter(c.content_type for c in chunk_docs))})"
        )
        return chunk_docs

    def chunk_content_fallback(self, content: str, metadata: Dict[str, Any]) -> List[ChunkData]:
        """
        Fallback chunking when Docling conversion fails (e.g. PyPDF2 text).

        Splits at '## Page N' boundaries, then enforces token budget via
        simple character-based splitting as a last resort.
        """
        self._ensure_initialized()
        file_name = metadata.get("file_name", "unknown")
        file_hash = metadata.get("file_hash", "")

        if not content.strip():
            return []

        page_pattern = re.compile(r'^## Page (\d+)', re.MULTILINE)
        page_matches = list(page_pattern.finditer(content))

        sections = []
        if page_matches:
            for idx, match in enumerate(page_matches):
                start = match.start()
                end = page_matches[idx + 1].start() if idx + 1 < len(page_matches) else len(content)
                page_no = int(match.group(1))
                text = content[start:end].strip()
                if text:
                    sections.append((text, [page_no]))
        else:
            sections = [(content.strip(), [])]

        import tiktoken
        enc = tiktoken.get_encoding(self.config.chunk_tokenizer)
        max_tokens = self.config.chunk_max_tokens

        chunk_docs: List[ChunkData] = []
        for text, page_nums in sections:
            tokens = enc.encode(text)
            start_tok = 0
            while start_tok < len(tokens):
                end_tok = min(start_tok + max_tokens, len(tokens))
                piece = enc.decode(tokens[start_tok:end_tok])
                if piece.strip():
                    chunk_hash = hashlib.sha256(piece.encode()).hexdigest()[:16]
                    i = len(chunk_docs)
                    meta = {
                        **metadata,
                        "content_type": "text",
                        "page_numbers": page_nums,
                        "page_no": page_nums[0] if page_nums else None,
                        "headings": [],
                        "heading_path": "",
                        "section_name": "",
                        "doc_title": "",
                        "chunk_hash": chunk_hash,
                        "chunk_size": len(piece),
                        "chunk_index": i,
                        "is_active": True,
                        "version": 1,
                    }
                    chunk_docs.append(ChunkData(
                        content=piece,
                        metadata=meta,
                        chunk_index=i,
                        file_name=file_name,
                        file_type=metadata.get("file_type", "unknown"),
                        content_type="text",
                        page_numbers=page_nums,
                        chunk_hash=chunk_hash,
                        embed_text=piece,
                        file_hash=file_hash,
                    ))
                start_tok = end_tok

        total = len(chunk_docs)
        for idx, cd in enumerate(chunk_docs):
            cd.chunk_index = idx
            cd.metadata["chunk_index"] = idx
            cd.metadata["total_chunks"] = total

        logger.info(f"Created {total} fallback chunks for {file_name}")
        return chunk_docs

    def _process_excel_streaming(self, file_path: str, metadata: Dict[str, Any]) -> List[ChunkData]:
        """
        Stream-read an Excel workbook and produce table-style chunks.

        Uses openpyxl in read-only mode to avoid loading full workbook into memory.
        Each sheet is treated as a separate 'table' and rows are grouped into
        chunks that respect the token budget configured by `chunk_max_tokens`.
        """
        # Avoid initializing Docling/tokenizers here to keep streaming
        # path lightweight and dependency-free for large workbooks.

        try:
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        except Exception as e:
            logger.error(f"Failed to open workbook for streaming: {e}")
            return []

        chunks: List[ChunkData] = []
        file_name = metadata.get("file_name", os.path.basename(file_path))
        file_hash = metadata.get("file_hash", "")
        max_tokens = self.config.chunk_max_tokens

        try:
            for s_idx, sheet_name in enumerate(wb.sheetnames):
                try:
                    ws = wb[sheet_name]
                except Exception:
                    continue

                rows = ws.iter_rows(values_only=True)
                headers = None
                header_line = f"Sheet: {sheet_name}"
                current_rows: List[str] = []

                for r in rows:
                    # Treat first non-empty row as header row
                    if headers is None:
                        if not any(c is not None and str(c).strip() for c in r):
                            continue
                        headers = ["" if c is None else str(c).strip() for c in r]
                        continue

                    cells = ["" if v is None else str(v).strip() for v in r]
                    if not any(cells):
                        continue

                    parts: List[str] = []
                    if headers and headers[0]:
                        parts.append(cells[0])
                    for h, val in zip((headers or [])[1:], cells[1:]):
                        if val:
                            parts.append(f"{h}: {val}" if h else str(val))

                    row_text = "Row: " + " | ".join(parts) if parts else "Row: " + " | ".join(cells)

                    # test token budget
                    trial = header_line + "\n" + "\n".join(current_rows + [row_text])
                    if current_rows and self._count_tokens(trial) > max_tokens:
                        piece = header_line + "\n" + "\n".join(current_rows)
                        chunk_hash = hashlib.sha256(piece.encode()).hexdigest()[:16]

                        if self.config.contextualize_embeddings:
                            embed_text = f"{sheet_name}\n\n" + piece
                        else:
                            embed_text = piece

                        idx = len(chunks)
                        meta = {
                            **metadata,
                            "content_type": "table",
                            "page_numbers": [s_idx + 1],
                            "page_no": s_idx + 1,
                            "headings": [],
                            "heading_path": "",
                            "section_name": sheet_name,
                            "doc_title": sheet_name,
                            "chunk_hash": chunk_hash,
                            "chunk_size": len(piece),
                            "chunk_index": idx,
                            "is_active": True,
                            "version": 1,
                        }

                        chunks.append(ChunkData(
                            content=piece,
                            metadata=meta,
                            chunk_index=idx,
                            file_name=file_name,
                            file_type=metadata.get("file_type", "xlsx"),
                            content_type="table",
                            page_numbers=[s_idx + 1],
                            headings=[],
                            heading_path="",
                            doc_title=sheet_name,
                            chunk_hash=chunk_hash,
                            embed_text=embed_text,
                            section_name=sheet_name,
                            file_hash=file_hash,
                        ))

                        current_rows = [row_text]
                    else:
                        current_rows.append(row_text)

                if current_rows:
                    piece = header_line + "\n" + "\n".join(current_rows)
                    chunk_hash = hashlib.sha256(piece.encode()).hexdigest()[:16]
                    if self.config.contextualize_embeddings:
                        embed_text = f"{sheet_name}\n\n" + piece
                    else:
                        embed_text = piece

                    idx = len(chunks)
                    meta = {
                        **metadata,
                        "content_type": "table",
                        "page_numbers": [s_idx + 1],
                        "page_no": s_idx + 1,
                        "headings": [],
                        "heading_path": "",
                        "section_name": sheet_name,
                        "doc_title": sheet_name,
                        "chunk_hash": chunk_hash,
                        "chunk_size": len(piece),
                        "chunk_index": idx,
                        "is_active": True,
                        "version": 1,
                    }

                    chunks.append(ChunkData(
                        content=piece,
                        metadata=meta,
                        chunk_index=idx,
                        file_name=file_name,
                        file_type=metadata.get("file_type", "xlsx"),
                        content_type="table",
                        page_numbers=[s_idx + 1],
                        headings=[],
                        heading_path="",
                        doc_title=sheet_name,
                        chunk_hash=chunk_hash,
                        embed_text=embed_text,
                        section_name=sheet_name,
                        file_hash=file_hash,
                    ))

            # final indexing
            total = len(chunks)
            for idx, cd in enumerate(chunks):
                cd.chunk_index = idx
                cd.metadata["chunk_index"] = idx
                cd.metadata["total_chunks"] = total

            wb.close()
            logger.info(f"Created {len(chunks)} streaming chunks for {file_name}")
            return chunks
        except Exception as e:
            logger.error(f"Error streaming workbook {file_path}: {e}")
            try:
                wb.close()
            except Exception:
                pass
            return []

    # ------------------------------------------------------------------
    # Document Processing
    # ------------------------------------------------------------------

    def _get_converter(
        self,
        enable_ocr: Optional[bool] = None,
        enable_table_extraction: Optional[bool] = None,
    ) -> "DocumentConverter":
        """
        Return the appropriate Docling converter.
        
        Args:
            enable_ocr: True=force OCR on, False=force OCR off, None=use global config
            enable_table_extraction: True=force on, False=force off, None=use global config
        """
        # If both are None, use the default converter (global config)
        if enable_ocr is None and enable_table_extraction is None:
            return self._converter
        
        # Resolve effective values
        ocr_val = enable_ocr if enable_ocr is not None else self.config.ocr_enabled
        table_val = enable_table_extraction if enable_table_extraction is not None else self.config.table_extraction_enabled
        
        cache_key = (ocr_val, table_val)
        if cache_key not in self._converter_cache:
            logger.info(f"Creating converter (on-demand): ocr={ocr_val}, table={table_val}")
            opts = self._PdfPipelineOptions(
                do_ocr=ocr_val,
                do_table_structure=table_val,
                ocr_options=self._ocr_options,
                generate_page_images=False,
                generate_picture_images=False,
                images_scale=1.0,
                force_backend_text=True,
            )
            fmt = self._PdfFormatOption(
                pipeline_cls=self._StandardPdfPipeline,
                backend=self._PyPdfiumDocumentBackend,
                pipeline_options=opts,
            )
            self._converter_cache[cache_key] = self._DocumentConverter(
                allowed_formats=[
                    self._InputFormat.PDF, self._InputFormat.DOCX,
                    self._InputFormat.PPTX, self._InputFormat.IMAGE,
                    self._InputFormat.XLSX,
                ],
                format_options={self._InputFormat.PDF: fmt},
            )
        return self._converter_cache[cache_key]

    def _ocr_pdf_with_tesseract(self, file_path: str) -> Optional[str]:
        """
        OCR a PDF using pypdfium2 (render) + pytesseract (OCR) directly.
        
        This bypasses Docling's layout model entirely. Used as a fallback
        when Docling fails (e.g. HuggingFace models unavailable offline).
        """
        try:
            import pytesseract
            import pypdfium2 as pdfium
            
            tesseract_path = os.path.expandvars(self.config.tesseract_path)
            pytesseract.pytesseract.tesseract_cmd = tesseract_path
            
            doc = pdfium.PdfDocument(file_path)
            pages_text = []
            
            for i in range(len(doc)):
                page = doc[i]
                # Render page to image at 300 DPI for good OCR quality
                bitmap = page.render(scale=300 / 72)
                pil_image = bitmap.to_pil()
                
                text = pytesseract.image_to_string(pil_image, lang="eng")
                if text and text.strip():
                    pages_text.append(f"## Page {i + 1}\n\n{text.strip()}")
                
                # Release resources
                pil_image.close()
            
            doc.close()
            
            if pages_text:
                content = "\n\n".join(pages_text)
                logger.info(
                    f"Tesseract OCR fallback successful: {len(pages_text)} pages, "
                    f"{len(content)} chars"
                )
                return content
            
            logger.warning("Tesseract OCR fallback produced no text")
            return None
            
        except Exception as e:
            logger.error(f"Tesseract OCR fallback failed: {e}")
            return None

    def _process_document_sync(
        self,
        file_path: str,
        parent_doc: Optional[str] = None,
        original_filename: Optional[str] = None,
        enable_ocr: Optional[bool] = None,
        enable_table_extraction: Optional[bool] = None,
    ) -> Optional[ProcessedDocument]:
        """
        Synchronous document processing (called from async methods).
        
        Args:
            file_path: Path to file to process
            parent_doc: Parent document if this is embedded
            original_filename: Original filename for metadata
            enable_ocr: Per-request OCR override (None=use global config)
            enable_table_extraction: Per-request table extraction override (None=use global config)
        """
        self._ensure_initialized()

        file_hash = self._get_file_hash_from_path(file_path)

        try:
            actual_filename = original_filename or os.path.basename(file_path)
            logger.info(f"Processing document: {actual_filename}")

            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)

            if file_size_mb > self.config.max_file_size_mb:
                logger.warning(f"File too large ({file_size_mb:.2f} MB): {actual_filename}")
                return ProcessedDocument(
                    content="",
                    metadata={},
                    chunks=[],
                    file_name=actual_filename,
                    file_hash=file_hash,
                    success=False,
                    error=f"File too large: {file_size_mb:.2f} MB",
                )

            file_ext = Path(actual_filename).suffix.lower()

            # Handle large Excel files with a streaming reader to avoid
            # loading the whole workbook into memory / Docling conversion.
            if file_ext == '.xlsx':
                metadata = {
                    "source": actual_filename,
                    "file_name": actual_filename,
                    "file_type": file_ext,
                    "file_size_mb": round(file_size_mb, 2),
                    "processed_date": datetime.now().isoformat(),
                    "parent_document": parent_doc,
                    "file_hash": file_hash,
                    "has_images": False,
                    "has_tables": True,
                    "has_embedded_files": False,
                }

                chunks = self._process_excel_streaming(file_path=file_path, metadata=metadata)
                return ProcessedDocument(
                    content="",
                    metadata=metadata,
                    chunks=chunks,
                    file_name=actual_filename,
                    file_hash=file_hash,
                    embedded_documents=[],
                    success=True,
                )

            # Convert .doc to .docx before processing
            converted_file = None
            if file_ext == '.doc':
                logger.info(f"Detected .doc file, converting to .docx: {actual_filename}")
                converted_file = self._convert_doc_to_docx(file_path)
                if converted_file:
                    file_path = converted_file
                    actual_filename = os.path.basename(file_path)
                    file_ext = Path(actual_filename).suffix.lower()
                    logger.info(f"Successfully converted to: {actual_filename}")
                else:
                    logger.error(f"Failed to convert .doc file: {actual_filename}")
                    return ProcessedDocument(
                        content="",
                        metadata={},
                        chunks=[],
                        file_name=actual_filename,
                        file_hash=file_hash,
                        success=False,
                        error="Failed to convert .doc file to .docx",
                    )

            metadata = {
                "source": actual_filename,
                "file_name": actual_filename,
                "file_type": file_ext,
                "file_size_mb": round(file_size_mb, 2),
                "processed_date": datetime.now().isoformat(),
                "parent_document": parent_doc,
                "file_hash": file_hash,
                "has_images": False,
                "has_tables": False,
                "has_embedded_files": False,
            }

            # Convert with Docling
            content = None
            docling_doc = None

            try:
                converter = self._get_converter(enable_ocr, enable_table_extraction)
                result = converter.convert(file_path)
                if hasattr(result, 'document') and result.document:
                    docling_doc = result.document
                    content = docling_doc.export_to_markdown()
            except Exception as docling_error:
                logger.warning(f"Docling conversion failed for {actual_filename}: {docling_error}")
                logger.info("Attempting fallback text extraction...")

            # Fallback 1: Use pypdf for PDFs if Docling fails (extracts embedded text)
            if content is None and file_path.lower().endswith('.pdf'):
                try:
                    import pypdf
                    with open(file_path, 'rb') as f:
                        reader = pypdf.PdfReader(f)
                        pages_text = []
                        for i, page in enumerate(reader.pages):
                            text = page.extract_text() or ""
                            if text.strip():
                                pages_text.append(f"## Page {i+1}\n\n{text}")
                        content = "\n\n".join(pages_text)
                        logger.info(f"Fallback PDF extraction successful: {len(pages_text)} pages")
                except Exception as pdf_error:
                    logger.error(f"Fallback PDF extraction failed: {pdf_error}")

            # Fallback 2: Tesseract OCR for scanned PDFs
            # If pypdf got very little text (< 100 chars = likely scanned/image PDF)
            # and OCR is enabled, use pytesseract + pypdfium2 directly
            ocr_active = enable_ocr if enable_ocr is not None else self.config.ocr_enabled
            if file_path.lower().endswith('.pdf') and ocr_active:
                if not content or len(content.strip()) < 100:
                    logger.info(
                        f"Scanned PDF detected ({len(content.strip()) if content else 0} chars from text extraction). "
                        f"Running Tesseract OCR fallback..."
                    )
                    ocr_content = self._ocr_pdf_with_tesseract(file_path)
                    if ocr_content and len(ocr_content.strip()) > len((content or "").strip()):
                        content = ocr_content

            if not content or not content.strip():
                logger.error(f"Failed to extract content from: {actual_filename}")
                return ProcessedDocument(
                    content="",
                    metadata={},
                    chunks=[],
                    file_name=actual_filename,
                    file_hash=file_hash,
                    success=False,
                    error="Failed to extract text content",
                )

            if docling_doc:
                if hasattr(docling_doc, 'pictures') and docling_doc.pictures:
                    metadata["has_images"] = True
                    metadata["num_images"] = len(docling_doc.pictures)
                if hasattr(docling_doc, 'tables') and docling_doc.tables:
                    metadata["has_tables"] = True
                    metadata["num_tables"] = len(docling_doc.tables)

            # Handle embedded DOCX files
            embedded_docs = []
            if file_path.endswith('.docx'):
                temp_dir = tempfile.mkdtemp()
                try:
                    embedded_files = self.extract_embedded_docx(file_path, temp_dir)
                    if embedded_files:
                        metadata["has_embedded_files"] = True
                        metadata["embedded_files"] = [os.path.basename(f) for f in embedded_files]
                        for embedded_file in embedded_files:
                            embedded_result = self._process_document_sync(embedded_file, parent_doc=file_path)
                            if embedded_result:
                                embedded_docs.append(embedded_result)
                finally:
                    shutil.rmtree(temp_dir, ignore_errors=True)

            # Chunk: HybridChunker if Docling succeeded, fallback otherwise
            # For very large documents (e.g. big XLSX with 1000+ rows), HybridChunker
            # can hang. Estimate token count and fall back to simpler chunker if too large.
            HYBRID_CHUNKER_TOKEN_LIMIT = 50_000
            if docling_doc:
                est_tokens = len(content) // 4 if content else 0
                if est_tokens > HYBRID_CHUNKER_TOKEN_LIMIT:
                    logger.warning(
                        f"Document too large for HybridChunker (~{est_tokens} tokens), "
                        f"using token-based fallback chunker for {actual_filename}"
                    )
                    chunks = self.chunk_content_fallback(content, metadata)
                else:
                    try:
                        import concurrent.futures
                        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
                            future = executor.submit(self.chunk_document, docling_doc, metadata)
                            chunks = future.result(timeout=120)
                    except concurrent.futures.TimeoutError:
                        logger.warning(
                            f"HybridChunker timed out (120s) for {actual_filename}, "
                            f"using token-based fallback chunker"
                        )
                        chunks = self.chunk_content_fallback(content, metadata)
            else:
                chunks = self.chunk_content_fallback(content, metadata)

            # Clean up converted .docx file if it exists
            if converted_file and os.path.exists(converted_file):
                try:
                    os.remove(converted_file)
                    logger.info(f"Cleaned up converted file: {converted_file}")
                except Exception as cleanup_err:
                    logger.warning(f"Failed to clean up converted file {converted_file}: {cleanup_err}")

            return ProcessedDocument(
                content=content,
                metadata=metadata,
                chunks=chunks,
                file_name=actual_filename,
                file_hash=file_hash,
                embedded_documents=embedded_docs,
                success=True,
            )

        except Exception as e:
            logger.error(f"Error processing {file_path}: {str(e)}")
            return ProcessedDocument(
                content="",
                metadata={},
                chunks=[],
                file_name=original_filename or os.path.basename(file_path),
                file_hash=file_hash,
                success=False,
                error=str(e),
            )

    async def process_file(
        self,
        file: UploadFile,
        enable_ocr: Optional[bool] = None,
        enable_table_extraction: Optional[bool] = None,
    ) -> ProcessedDocument:
        """
        Process a single uploaded file.
        
        Args:
            file: FastAPI UploadFile object
            enable_ocr: Per-request OCR override (None=use global config)
            enable_table_extraction: Per-request table extraction override (None=use global config)
            
        Returns:
            ProcessedDocument with content, chunks, and metadata
        """
        validation = self.validate_file(file)
        if not validation.valid:
            return ProcessedDocument(
                content="",
                metadata={},
                chunks=[],
                file_name=file.filename or "unknown",
                file_hash="",
                success=False,
                error=validation.message,
            )

        temp_dir = tempfile.mkdtemp()
        try:
            file_path = await self.save_upload_file(file, temp_dir)
            result = self._process_document_sync(
                file_path, original_filename=file.filename,
                enable_ocr=enable_ocr, enable_table_extraction=enable_table_extraction,
            )
            return result or ProcessedDocument(
                content="",
                metadata={},
                chunks=[],
                file_name=file.filename or "unknown",
                file_hash="",
                success=False,
                error="Processing returned no result",
            )
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    async def process_files(self, files: List[UploadFile]) -> AsyncGenerator[ProcessingStatus, None]:
        """Process multiple files with progress updates."""
        total_files = len(files)

        for i, file in enumerate(files):
            file_name = file.filename or f"file_{i}"

            yield ProcessingStatus(
                file_name=file_name,
                status="starting",
                progress=(i / total_files),
                message=f"Starting processing of {file_name}",
            )

            validation = self.validate_file(file)
            if not validation.valid:
                yield ProcessingStatus(
                    file_name=file_name,
                    status="error",
                    progress=((i + 1) / total_files),
                    message=validation.message,
                    error=validation.message,
                )
                continue

            yield ProcessingStatus(
                file_name=file_name,
                status="processing",
                progress=(i / total_files) + (0.5 / total_files),
                message=f"Converting {file_name}...",
            )

            try:
                result = await self.process_file(file)

                if result.success:
                    yield ProcessingStatus(
                        file_name=file_name,
                        status="complete",
                        progress=((i + 1) / total_files),
                        message=f"Completed {file_name}",
                        chunks_created=len(result.chunks),
                    )
                else:
                    yield ProcessingStatus(
                        file_name=file_name,
                        status="error",
                        progress=((i + 1) / total_files),
                        message=result.error or "Unknown error",
                        error=result.error,
                    )
            except Exception as e:
                logger.error(f"Error processing {file_name}: {e}")
                yield ProcessingStatus(
                    file_name=file_name,
                    status="error",
                    progress=((i + 1) / total_files),
                    message=str(e),
                    error=str(e),
                )

    def clear_processed_cache(self) -> None:
        """Reset lazy-loaded state so next call re-initializes."""
        self._initialized = False
        self._converter = None
        self._converter_cache = {}
        self._chunker = None
        logger.debug("DocumentService cache cleared")


# ============================================================================
# Singleton Factory
# ============================================================================

_document_service: Optional[DocumentService] = None


def get_document_service() -> DocumentService:
    """Get or create the document service singleton."""
    global _document_service
    if _document_service is None:
        _document_service = DocumentService()
    return _document_service
